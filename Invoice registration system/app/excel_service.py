
import os
from openpyxl import Workbook, load_workbook

HEADERS = ["收票日期","交票人","PDF文档名称","开票日期","发票号码","购买方信息名称","销售方信息名称","金额","项目名称","开票类型","状态"]
BASIC_EXPORT_HEADERS = ["收票日期","交票人","PDF 文档名称","开票日期","发票号码","购买方信息名称","销售方信息名称","金额","项目名称"]
ENTRY_EXPORT_HEADERS = ["收票日期","交票人","PDF 文档名称","开票日期","发票号码","购买方信息名称","销售方信息名称","金额","项目名称","备注"]
REIMBURSE_LOG_HEADERS = ["核销日期","员工","工号","核销金额","核销事项","备注","收票日期","交票人","PDF 文档名称","开票日期","发票号码","购买方信息名称","销售方信息名称","项目名称"]

def _ensure_wb(path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        ws.append(HEADERS)
        wb.save(path)

def append_invoice_to_excel(path, row):
    _ensure_wb(path)
    wb = load_workbook(path)
    ws = wb.active
    ws.append([
        row.get("receipt_date"), row.get("submitter_name"), row.get("pdf_filename"),
        row.get("invoice_date"), row.get("invoice_no"), row.get("buyer_name"),
        row.get("seller_name"), row.get("amount"), row.get("project_name"),
        row.get("invoice_type_name"), row.get("status")
    ])
    wb.save(path)

def export_invoice_rows_to_excel(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    headers = ["ID","员工","工号","收票日期","交票人","文件名","开票日期","发票号","购买方","销售方","金额","项目","开票类型","状态","财务备注","提交时间"]
    ws.append(headers)
    for r in rows:
        ws.append([
            r.get("id"), r.get("employee_name"), r.get("employee_no"),
            str(r.get("receipt_date") or ""), r.get("submitter_name"), r.get("pdf_filename"),
            str(r.get("invoice_date") or ""), r.get("invoice_no"), r.get("buyer_name"),
            r.get("seller_name"), float(r.get("amount") or 0), r.get("project_name"),
            r.get("type_name"), r.get("status"), r.get("finance_comment"), str(r.get("created_at") or "")
        ])
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)

def export_invoice_basic_rows_to_excel(path, rows, sheet_title="Invoices"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(BASIC_EXPORT_HEADERS)
    for r in rows:
        ws.append([
            str(r.get("receipt_date") or ""),
            r.get("submitter_name"),
            r.get("pdf_filename"),
            str(r.get("invoice_date") or ""),
            r.get("invoice_no"),
            r.get("buyer_name"),
            r.get("seller_name"),
            float(r.get("amount") or 0),
            r.get("project_name"),
        ])
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)

def export_entry_rows_to_excel(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "录入明细"
    ws.append(ENTRY_EXPORT_HEADERS)
    for r in rows:
        ws.append([
            str(r.get("receipt_date") or ""),
            r.get("submitter_name"),
            r.get("pdf_filename"),
            str(r.get("invoice_date") or ""),
            r.get("invoice_no"),
            r.get("buyer_name"),
            r.get("seller_name"),
            float(r.get("amount") or 0),
            r.get("project_name"),
            "核销未通过" if r.get("status") == "Rejected" else (r.get("finance_comment") or ""),
        ])
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)

def export_reimbursement_logs_to_excel(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "核销明细"
    ws.append(REIMBURSE_LOG_HEADERS)
    for r in rows:
        ws.append([
            str(r.get("reimbursed_at") or ""),
            r.get("employee_name"),
            r.get("employee_no"),
            float(r.get("reimburse_amount") or 0),
            r.get("reimburse_item_name"),
            r.get("comment") or "",
            str(r.get("receipt_date") or ""),
            r.get("submitter_name"),
            r.get("pdf_filename"),
            str(r.get("invoice_date") or ""),
            r.get("invoice_no"),
            r.get("buyer_name"),
            r.get("seller_name"),
            r.get("project_name"),
        ])
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
